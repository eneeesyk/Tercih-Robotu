import pyexcel as pe

import  xlsxwriter
import openpyxl

import excel_writer


def unique(list1):
    # initialize a null list
    unique_list = []

    # traverse for all elements
    for x in list1:
        # check if exists in unique_list or not
        if x not in unique_list:
            unique_list.append(x)
    return unique_list


import pandas as pd

FACULTY_LIST = []
UNI_NAME = []
FACULTY_NAME = []
FACULTY_NAME_UNIQUE = []
UNIVERSITY_LIST = []
PROGRAM_CODE = []
PROGRAM_NAME = []
TABAN_PUANI = []
BIRINCI_KONT = []
KONTENJAN = []
BASARI_SIRASI = []
PUAN_TURU = []
OGR_SURE = []
DESC = []


def has_numbers(inputString):
    return any(char.isdigit() for char in inputString)


filename = "2021_4.xls"
records = pe.get_array(file_name=filename)

for i in range(len(records)):
    if len(records[i][0]) == 0:
        if "Fakülte" in records[i][1]:
            if not records[i][1] in FACULTY_LIST:
                FACULTY_LIST.append(records[i][1])  # ünik yapıyor
        if "ÜNİVERSİTE" in records[i][1]:
                UNIVERSITY_LIST.append(records[i][1])  # ünik yapıyor


    else:
        OGR_SURE.append(records[i][2])
        PUAN_TURU.append(records[i][3])
        KONTENJAN.append(records[i][4])
        BIRINCI_KONT.append(records[i][5])
        DESC.append(records[i][6])
        BASARI_SIRASI.append(records[i][8])
        TABAN_PUANI.append(records[i][9])
# TODO FACULTY_LIST
for i in range(len(records)):
    if len(records[i][0]) == 0:
        if records[i][1] in UNIVERSITY_LIST:
            uni = records[i][1]
        elif records[i][1] in FACULTY_LIST:
            if 'uni' in vars():
                UNI_NAME.append(uni)
            FACULTY_NAME.append(records[i][1])

OGR_SURE = [int(i) for i in OGR_SURE[2:]]
PUAN_TURU = PUAN_TURU[2:]
FACULTY_LIST = FACULTY_LIST[1:]
KONTENJAN = [int(i) for i in KONTENJAN[2:]]
# BIRINCI_KONT = [int(i) for i in BIRINCI_KONT[2:]]
DESC = DESC[2:]
# BASARI_SIRASI = [int(i) for i in BASARI_SIRASI[2:]]
# TABAN_PUANI = [float(i) for i in TABAN_PUANI[2:]]


excel_writer.excel_write("university.xlsx", UNIVERSITY_LIST, title="uni_id", title2="uni_name",column_2=1)
for a in range(len(FACULTY_NAME)):
    if "Fakülte" in FACULTY_NAME[a]:
        FACULTY_NAME_UNIQUE.append(FACULTY_NAME[a])

FACULTY_NAME_UNIQUE = unique(FACULTY_NAME_UNIQUE)
excel_writer.excel_write("faculty_name.xlsx", FACULTY_NAME_UNIQUE, title="faculty_name_id", title2="faculty_name", column_2=1)
excel_writer.excel_write("uni_faculty_name.xlsx", UNI_NAME, title="uni_name", title2="faculty_name", content2=FACULTY_NAME, column_2=1)
excel_writer.excel_write("uni_fac_id.xlsx", FACULTY_NAME_UNIQUE, title="faculty_name_id", title2="faculty_name", title3="uni_id", title4="uni_name", content2=UNI_NAME,column_2=3)

#excel_writer.merge()
# excel_writer.excel_write_fk("faculty.xls", id, id2, title="faculty_id", title2="uni_id")


# TODO EXCELI YAZ

deneme1=openpyxl.load_workbook("uni_faculty_name.xlsx")  #deneme1=uni_fac_name
deneme2=openpyxl.load_workbook("uni_fac_id.xlsx")           #deneme2=uni_fac_id

deneme1_sheet=deneme1['Sheet1']
deneme2_sheet=deneme2['Sheet1']

for i in deneme1_sheet.iter_rows():
    uni_name=i[0].value
    fac_name=i[1].value
    row_number=i[0].row
    #print(i[1].value)
    row_number2=i[1].row

    for j in deneme2_sheet.iter_rows():
        if j[0].value== uni_name:
            deneme1_sheet.cell(row=row_number,column=3).value=j[1].value
            #print(j[1].value)
        #ELIF KULLANINCA OLMUYOR
        #elif j[2].value==fac_name:
        #    deneme1_sheet.cell(row=row_number, column=4).value = j[3].value
    for j in deneme2_sheet.iter_rows():
        if j[2].value==fac_name:
            deneme1_sheet.cell(row=row_number, column=4).value = j[3].value
            #print(j[3].value)

deneme1.save("deneme3.xlsx")
deneme3=openpyxl.load_workbook("deneme3.xlsx")
deneme3_sheet=deneme3['Sheet1']
deneme3_sheet.delete_cols(idx=1,amount=2)

deneme3.save("faculty.xlsx")
#worksheet.write_formula('A2','=VLOOKUP("İstanbul Medipol Üni ","H1:I1",2,FALSE')
##ws['A2']="=VLOOKUP(C2;H:I;2;0)"
#ws['A2']="=SUM(I2:I225)"


##wb.save("exel.xlsx")
# TODO EXCELI YAZ